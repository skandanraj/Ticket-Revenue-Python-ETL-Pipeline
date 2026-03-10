"""
Match ticket data with revenue records.

Logic
-----
1. Load ticket and revenue datasets
2. Combine multiple tests under the same ticket
3. Match revenue using:
      - UHID (preferred)
      - Phone number (fallback)
4. Ensure invoice date is after ticket creation
5. Export results to Excel with validation checks

Output
------
Excel file with two sheets:
- Raw Data
- Validation
"""

import pandas as pd
from openpyxl import load_workbook


# file paths
ticket_input = r"input folder path/Tickets.xlsx"
revenue_input = r"input folder path/Revenue.xlsx"
final_output = r"output folder path/App_Tickets_With_Revenue.xlsx"


# -----------------------------
# Load ticket data
# -----------------------------
tickets = pd.read_excel(ticket_input)

# clean column names
tickets.columns = tickets.columns.str.strip()

# convert created time to datetime
tickets["Created At"] = pd.to_datetime(tickets["Created At"])

# keep only required columns
tickets = tickets[
    [
        "Ticket Id",
        "Contact Name",
        "UHID",
        "Phone No",
        "Branch",
        "Source",
        "Priority",
        "Status",
        "Created At",
        "test_id",
        "test_name",
    ]
]


# -----------------------------
# Combine tests under same ticket
# -----------------------------
tickets = (
    tickets.groupby("Ticket Id", as_index=False)
    .agg(
        {
            "Contact Name": "first",
            "UHID": "first",
            "Phone No": "first",
            "Branch": "first",
            "Source": "first",
            "Priority": "first",
            "Status": "first",
            "Created At": "first",
            "test_id": lambda x: ", ".join(x.astype(str).unique()),
            "test_name": lambda x: ", ".join(x.astype(str).unique()),
        }
    )
)

print("Tickets loaded:", len(tickets))


# -----------------------------
# Load revenue data
# -----------------------------
revenue = pd.read_excel(revenue_input)

revenue.columns = revenue.columns.str.strip()

# convert invoice date
revenue["InvoiceDate"] = pd.to_datetime(revenue["InvoiceDate"])


# -----------------------------
# Revenue matching logic
# -----------------------------
def match_revenue(ticket):

    uhid = ticket["UHID"]
    phone = ticket["Phone No"]
    branch = ticket["Branch"]
    created = ticket["Created At"]

    row = ticket.to_dict()

    # try matching using UHID first
    if pd.notna(uhid) and str(uhid).strip() != "":

        candidates = revenue[
            (revenue["RegistrationNo"] == uhid)
            & (revenue["Source_Branch"] == branch)
            & (revenue["InvoiceDate"] >= created)
        ]

        match_type = "UHID"

    # fallback to phone number
    else:

        candidates = revenue[
            (revenue["Phone No"] == phone)
            & (revenue["Source_Branch"] == branch)
            & (revenue["InvoiceDate"] >= created)
        ]

        match_type = "PHONE"

    # if revenue record found
    if not candidates.empty:

        # pick earliest invoice
        best = candidates.sort_values("InvoiceDate").iloc[0]

        row.update(
            {
                "InvoiceNo": best.get("InvoiceNo"),
                "RegistrationNo": best.get("RegistrationNo"),
                "Revenue Phone No": best.get("Phone No"),
                "PatientName": best.get("PatientName"),
                "Source_Branch": best.get("Source_Branch"),
                "InvoiceDate": best.get("InvoiceDate"),
                "Service Name": best.get("Service Name"),
                "Gross Amount": best.get("Gross Amount"),
                "Match_Type": match_type,
            }
        )

    # no match found
    else:

        row.update(
            {
                "InvoiceNo": pd.NA,
                "RegistrationNo": pd.NA,
                "Revenue Phone No": pd.NA,
                "PatientName": pd.NA,
                "Source_Branch": pd.NA,
                "InvoiceDate": pd.NA,
                "Service Name": pd.NA,
                "Gross Amount": pd.NA,
                "Match_Type": "NO MATCH",
            }
        )

    return pd.Series(row)


# -----------------------------
# run matching for all tickets
# -----------------------------
final_df = tickets.apply(match_revenue, axis=1)

# rename ticket phone column for clarity
final_df = final_df.rename(columns={"Phone No": "Ticket Phone No"})


# -----------------------------
# Raw data sheet
# -----------------------------
raw_df = final_df[
    [
        "Ticket Id",
        "Contact Name",
        "UHID",
        "Ticket Phone No",
        "Branch",
        "Priority",
        "Status",
        "test_id",
        "test_name",
        "Created At",
        "InvoiceNo",
        "RegistrationNo",
        "Revenue Phone No",
        "PatientName",
        "Source_Branch",
        "InvoiceDate",
        "Service Name",
        "Gross Amount",
        "Match_Type",
    ]
]


# -----------------------------
# Validation sheet
# -----------------------------
validation_df = raw_df[
    [
        "Ticket Id",
        "UHID",
        "RegistrationNo",
        "Ticket Phone No",
        "Revenue Phone No",
        "Branch",
        "Source_Branch",
        "Created At",
        "InvoiceDate",
        "Gross Amount",
    ]
]


# -----------------------------
# Write Excel output
# -----------------------------
with pd.ExcelWriter(final_output, engine="openpyxl") as writer:
    raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
    validation_df.to_excel(writer, sheet_name="Validation", index=False)


# -----------------------------
# Add validation formulas
# -----------------------------
wb = load_workbook(final_output)
ws = wb["Validation"]

ws["K1"] = "Test UHID"
ws["L1"] = "Test Phone No"
ws["M1"] = "Test Branch"
ws["N1"] = "Test Date"

for row in range(2, ws.max_row + 1):

    ws[f"K{row}"] = f'=IF(B{row}=C{row},"YES","NO")'
    ws[f"L{row}"] = f'=IF(D{row}=E{row},"YES","NO")'
    ws[f"M{row}"] = f'=IF(F{row}=G{row},"YES","NO")'
    ws[f"N{row}"] = f'=IF(H{row}<=I{row},"YES","NO")'


wb.save(final_output)


print("Script finished")
print("Rows processed:", len(raw_df))
print("Output file:", final_output)
